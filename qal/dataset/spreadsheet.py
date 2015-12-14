"""
Created on Dec 28, 2013

@author: Nicklas Boerjesson
"""
from qal.common.strings import make_path_absolute, string_to_bool

from qal.dataset.custom import CustomDataset
from qal.common.discover import import_error_to_help


def none_to_zero(_value):
    if _value is None or _value is False:
        return 0
    else:
        return _value


class SpreadsheetDataset(CustomDataset):
    """The Spreadsheet dataset can read data from a spreadsheet, currently only Excel files, and store that data in its.
    """

    has_header = None
    """True if data begins with a header row containing field names."""
    filename = None
    """The name of the file."""
    field_names = None
    """The names of the fields(if applicable, see has_header)."""
    sheet_name = None
    """The name of the worksheet to use, typically "Sheet1"."""
    x_offset = None
    """The leftmost column where data is held"""
    y_offset = None
    """The topmost row where data is held, *including the header row*."""

    def read_resource_settings(self, _resource):
        self._base_path = _resource.base_path

        if _resource.type.upper() != 'SPREADSHEET':
            raise Exception(
                "SpreadsheetDataset.read_resource_settings.parse_resource error: Wrong resource type: " +
                _resource.type)
        self.filename = _resource.filename

        if _resource.has_header:
            self.has_header = string_to_bool(str(_resource.has_header))
        else:
            self.has_header = None

        self.sheet_name = _resource.sheet_name

        if hasattr(_resource, "x_offset") and _resource.x_offset is not None:
            self.x_offset = int(_resource.x_offset)
        else:
            self.x_offset = 0
        if hasattr(_resource, "y_offset") and _resource.y_offset is not None:
            self.y_offset = int(_resource.y_offset)
        else:
            self.y_offset = 0

    def write_resource_settings(self, _resource):
        _resource.type = 'SPREADSHEET'
        _resource.data.clear()
        _resource.filename = self.filename
        _resource.has_header = self.has_header
        _resource.sheet_name = self.sheet_name
        _resource.x_offset = self.x_offset
        _resource.y_offset = self.y_offset

    def __init__(self, _filename=None, _has_header=None, _resource=None, _sheet_name=None, _x_offset=None,
                 _y_offset=None):
        """
        Constructor
        """
        super(SpreadsheetDataset, self).__init__()
        if _resource is not None:
            self.read_resource_settings(_resource)
        else:

            if _filename is not None:
                self.filename = _filename
            else:
                self.filename = None
            if _has_header is not None:
                self.has_header = _has_header
            else:
                self.has_header = None
            if _sheet_name is not None:
                self.sheet_name = _sheet_name
            else:
                self.sheet_name = None

            if _x_offset is not None:
                self.x_offset = _x_offset
            else:
                self.x_offset = None
            if _y_offset is not None:
                self.y_offset = _y_offset
            else:
                self.y_offset = None

    @staticmethod
    def load_xls(_filename, _sheet_name, _x_offset, _y_offset, _header_offset):
        print("SpreadsheetDataset.load_xls: using xlrd")

        try:
            import xlrd
        except ImportError as _err:
            raise Exception(import_error_to_help(_module="xlrd", _err_obj=_err, _pip_package="xlrd",
                                                 _apt_package=None, _win_package=None))
        try:
            _workbook = xlrd.open_workbook(_filename)
            _sheet = _workbook.sheet_by_name(_sheet_name)

            if _header_offset > 0:
                _field_names = _sheet.row_values(rowx=_y_offset, start_colx=_x_offset, end_colx=_sheet.ncols)
            else:
                _field_names = list("Column_" + str(x) for x in range(_y_offset, _sheet.ncols))

            _data_table = []

            for _curr_y in range(_y_offset + _header_offset, _sheet.nrows):
                _data_table.append(_sheet.row_values(_curr_y, start_colx=_x_offset, end_colx=_sheet.ncols))

            return _data_table, _field_names

        except IOError as e:
            raise Exception("SpreadsheetDataset.load: Error reading file:" + str(e))

    @staticmethod
    def load_xlsx(_filename, _sheet_name, _x_offset, _y_offset, _header_offset):
        print("SpreadsheetDataset.load_xlsx: using openpyxl")
        try:
            from openpyxl import load_workbook
        except ImportError as _err:
            raise Exception(import_error_to_help(_module="openpyxl",
                                                 _err_obj=_err,
                                                 _pip_package="openpyxl",
                                                 _apt_package=None,
                                                 _win_package=None))
        try:
            _workbook = load_workbook(_filename)
            _sheet = _workbook.active
            _number_of_columns = _sheet.get_highest_column()
            _number_of_rows = _sheet.get_highest_row()
            _data_table = []
            for _curr_row in _sheet.iter_rows():
                _data_table.append([_cell.value for _cell in _curr_row])

            if _header_offset > 0:
                _field_names = [_sheet.cell(column=_x, row=1 + _y_offset).value for _x in
                                range(1 + _x_offset, _number_of_columns)]
            else:
                _field_names = list("Column_" + str(x) for x in range(1 + _x_offset, _number_of_columns))

            for _curr_y in range(1 + _y_offset + _header_offset, _number_of_rows):
                _data_table.append(
                    [_sheet.cell(column=_x, row=_curr_y).value for _x in range(1 + _x_offset, _number_of_columns)])

            return _data_table, _field_names

        except IOError as e:
            raise Exception("SpreadsheetDataset.load: Error reading file:" + str(e))

    def load(self):
        """Load data. """

        import os.path

        _extension = os.path.splitext(self.filename)[1]

        _x_offset = none_to_zero(self.x_offset)
        _y_offset = none_to_zero(self.y_offset)

        if self.has_header:
            _header_offset = 1
        else:
            _header_offset = 0

        _filename = make_path_absolute(self.filename, self._base_path)

        if _extension.lower() in [".xls", ".xlsx"]:
            self.data_table, self.field_names = self.load_xls(_filename=_filename, _sheet_name=self.sheet_name,
                                                              _x_offset=_x_offset, _y_offset=_y_offset,
                                                              _header_offset=_header_offset)
        # elif _extension.lower() in [".xlsx"]:
        # self.data_table, self.field_names = self.load_xlsx(_filename=_filename, _sheet_name=self.sheet_name,
        # _x_offset=_x_offset, _y_offset=_y_offset,
        #                                    _header_offset = _header_offset)

        elif _extension.lower() == ".odt":
            # from ezodf import newdoc, Paragraph, Heading, Sheet
            raise Exception("SpreadsheetDataset.load: Open Document Spreadsheet not implemented yet.")

        else:
            raise Exception("SpreadsheetDataset.load: Unsupported file type \"" + _extension + "\"")

    def save_xlsx(self, _filename, _data_table, _number_of_rows, _number_of_columns,
                  _x_offset, _y_offset, _header_offset, _field_names, _sheet_name):
        print("SpreadsheetDataset.save_xlsx: using openpyxl")

        try:
            import openpyxl
        except ImportError as _err:
            raise Exception(import_error_to_help(_module="openpyxl",
                                                 _err_obj=_err,
                                                 _pip_package="openpyxl",
                                                 _apt_package=None,
                                                 _win_package=None))

        try:

            _workbook = openpyxl.Workbook()
            _sheet = _workbook.create_sheet(title=_sheet_name)

            if _header_offset > 0:
                for _curr_col_idx in range(0, _number_of_columns):
                    _sheet.cell(row=1 + self.y_offset, column=1 + _curr_col_idx + _x_offset).value = _field_names[
                        _curr_col_idx]

            for _curr_row_idx in range(0, _number_of_rows - 1):
                for _curr_col_idx in range(0, _number_of_columns):
                    _sheet.cell(row=1 + _curr_row_idx + _y_offset + _header_offset,
                                column=1 + _curr_col_idx + _x_offset).value = \
                        _data_table[_curr_row_idx][_curr_col_idx]

            _workbook.save(filename=_filename)
        except IOError as e:
            raise Exception("SpreadsheetDataset.save_xlsx: Error saving file:" + str(e))

    def save_xls(self, _filename, _data_table, _number_of_rows, _number_of_columns,
                 _x_offset, _y_offset, _header_offset, _field_names, _sheet_name):
        print("SpreadsheetDataset.save_xls: using xlwt")

        try:
            import xlwt
        except ImportError as _err:
            raise Exception(import_error_to_help(_module="xlwt",
                                                 _err_obj=_err,
                                                 _pip_package="xlwt",
                                                 _apt_package=None,
                                                 _win_package=None))
        try:

            _workbook = xlwt.Workbook()
            _sheet = _workbook.add_sheet(sheetname=_sheet_name, cell_overwrite_ok=True)
            # Write headers
            if _header_offset > 0:
                for _curr_col_idx in range(0, _number_of_columns):
                    _sheet.write(self.y_offset, _curr_col_idx + _x_offset, label=_field_names[_curr_col_idx])
            # Write data
            for _curr_row_idx in range(0, _number_of_rows - 1):
                for _curr_col_idx in range(0, _number_of_columns):
                    _sheet.write(_curr_row_idx + _y_offset + _header_offset, _curr_col_idx + _x_offset,
                                 label=_data_table[_curr_row_idx][_curr_col_idx])

            _workbook.save(filename_or_stream=_filename)
        except IOError as e:
            raise Exception("SpreadsheetDataset.save_xls: Error saving file:" + str(e))

    def save(self, _save_as=None):
        """Save data"""
        _number_of_rows = len(self.data_table)
        if _number_of_rows > 0:
            _number_of_columns = len(self.data_table[0])
        print("SpreadsheetDataset.save: Filename='" + str(self.filename) + "'")
        import os.path

        _extension = os.path.splitext(self.filename)[1]
        if _save_as:
            _filename = _save_as
        else:
            _filename = make_path_absolute(self.filename, self._base_path)

        if self.y_offset is None:
            self.y_offset = 0
        if self.has_header:
            _header_offset = 1
        else:
            _header_offset = 0

        if _extension.lower() == ".xls":
            self.save_xls(_filename, _data_table=self.data_table, _number_of_rows=_number_of_rows,
                          _number_of_columns=_number_of_columns, _x_offset=self.x_offset,
                          _y_offset=self.y_offset, _header_offset=_header_offset, _field_names=self.field_names,
                          _sheet_name=self.sheet_name)

        elif _extension.lower() == ".xlsx":
            self.save_xlsx(_filename, _data_table=self.data_table, _number_of_rows=_number_of_rows,
                           _number_of_columns=_number_of_columns, _x_offset=self.x_offset,
                           _y_offset=self.y_offset, _header_offset=_header_offset, _field_names=self.field_names,
                           _sheet_name=self.sheet_name)

        elif _extension.lower() == ".odt":
            # from ezodf import newdoc, Paragraph, Heading, Sheet
            raise Exception("SpreadsheetDataset.load: Open Document Spreadsheet not implemented yet.")

        else:
            raise Exception("SpreadsheetDataset.load: Unsupported file type \"" + _extension + "\"")
